import socket
import random
from openpyxl import load_workbook

def read_plate_codes(filename):
    wb = load_workbook(filename)
    ws = wb.active
    city_codes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        city, code = row
        city_codes[city.strip()] = int(code)
    return city_codes

def start_server(host, port, filename):
    city_codes = read_plate_codes(filename)
    cities = list(city_codes.keys())
    
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((host, port))
        s.listen()
        client = 1

        while True:
            print(f"Waiting for {client}th client connection")
            print("Server waiting for connection...")
            conn, addr = s.accept()
            with conn:
                print("Client connected from:", addr)
                chosen_city = random.choice(cities)
                msg = f"What is the plate code of {chosen_city}\n"
                conn.sendall(msg.encode())
                
                while True:
                    data = conn.recv(1024).decode().strip()
                    print("Received from client:", data)
                    
                    if data.isdigit():
                        code = int(data)
                        if code < 1 or code > 81:
                            msg = "Number exceeds the range. Game over.\n"
                            conn.sendall(msg.encode())
                            break  
                        elif city_codes[chosen_city] == code:
                            msg = "Correct!\n"
                            conn.sendall(msg.encode())
                            break  
                        else:
                            city_for_code = next((city for city, c in city_codes.items() if c == code), None)
                            if city_for_code:
                                msg = f"You have entered the plate code of {city_for_code}\n"
                            else:
                                msg = "No city found with this plate code.\n"
                    else:
                        msg = "You entered a non-numeric value. Game over.\n"
                        conn.sendall(msg.encode())
                        break  

                    if data.upper() == "END":
                        msg = "END\n"
                        conn.sendall(msg.encode())
                        break  

                    conn.sendall(msg.encode())
            client = client + 1
            if data.upper() == "END":
                break

if __name__ == "__main__":
    HOST = 'localhost'
    PORT = 65435
    FILENAME = 'plate_list.xlsx'
    start_server(HOST, PORT, FILENAME)
